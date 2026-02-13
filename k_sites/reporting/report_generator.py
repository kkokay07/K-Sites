"""
Report Generator for K-Sites

This module generates publication-ready HTML reports from K-Sites pipeline output.
Includes comprehensive CSV, Excel, FASTA, and GenBank export functionality.
"""

import json
import html
import csv
import os
from typing import Dict, Any, List
from io import StringIO
import logging

# Set up logging
logger = logging.getLogger(__name__)


def generate_html_report(pipeline_output: dict, output_path: str) -> None:
    """
    Generate a publication-ready HTML report from pipeline output.
    
    Args:
        pipeline_output: Output dictionary from run_k_sites_pipeline
        output_path: Path to save the HTML report
    """
    logger.info(f"Generating HTML report at {output_path}")
    
    try:
        # Extract data from pipeline output
        metadata = pipeline_output.get("metadata", {})
        genes = pipeline_output.get("genes", [])
        
        # Count statistics
        total_genes_screened = len(genes)
        total_guides = sum(len(gene.get("guides", [])) for gene in genes)
        pathway_conflict_guides = sum(
            1 for gene in genes 
            for guide in gene.get("guides", []) 
            if guide.get("pathway_conflict", False)
        )
        
        # Generate the HTML report
        html_content = _generate_report_html(
            metadata, 
            genes, 
            total_genes_screened, 
            total_guides, 
            pathway_conflict_guides,
            output_path
        )
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        # Generate additional output formats (CSV, Excel, FASTA)
        _generate_additional_outputs(pipeline_output, output_path)
        
        logger.info(f"Successfully generated HTML report at {output_path}")
        
    except Exception as e:
        logger.error(f"Failed to generate HTML report: {str(e)}")
        raise


def _generate_additional_outputs(pipeline_output: dict, output_path: str) -> None:
    """
    Generate additional output formats alongside the HTML report.
    
    Args:
        pipeline_output: Output dictionary from run_k_sites_pipeline
        output_path: Path to the HTML report (used to derive other paths)
    """
    try:
        base_path = output_path.replace('.html', '')
        
        # Generate comprehensive CSV
        csv_path = f"{base_path}_comprehensive.csv"
        _generate_csv_report(pipeline_output, csv_path)
        
        # Generate gene summary CSV
        summary_csv_path = f"{base_path}_gene_summary.csv"
        _generate_gene_summary_csv(pipeline_output, summary_csv_path)
        
        # Generate Excel report
        excel_path = f"{base_path}_comprehensive.xlsx"
        _generate_excel_report(pipeline_output, excel_path)
        
        # Generate FASTA file
        fasta_path = f"{base_path}_grna_sequences.fasta"
        _generate_fasta_report(pipeline_output, fasta_path)
        
        # Generate GenBank file if available
        try:
            from k_sites.reporting.genbank_export import generate_genbank_from_pipeline
            genbank_path = f"{base_path}_sequences.gb"
            generate_genbank_from_pipeline(pipeline_output, genbank_path)
        except ImportError:
            logger.warning("GenBank export not available, skipping")
        
        logger.info(f"Generated all additional output formats")
        
    except Exception as e:
        logger.warning(f"Some additional outputs could not be generated: {str(e)}")


def _generate_csv_report(pipeline_output: dict, csv_path: str) -> None:
    """
    Generate a detailed CSV report with all results.
    
    Args:
        pipeline_output: Output dictionary from run_k_sites_pipeline
        csv_path: Path to save the CSV report
    """
    try:
        genes = pipeline_output.get("genes", [])
        metadata = pipeline_output.get("metadata", {})
        
        # Prepare CSV data
        rows = []
        
        # Add metadata as header comments
        rows.append([f"# K-Sites CRISPR Design Report"])
        rows.append([f"# GO Term: {metadata.get('go_term', 'N/A')}"])
        rows.append([f"# Organism: {metadata.get('organism', 'N/A')}"])
        rows.append([f"# Timestamp: {metadata.get('timestamp', 'N/A')}"])
        rows.append([f"# Evidence Filter: {metadata.get('evidence_filter', 'N/A')}"])
        rows.append([f"# Max Pleiotropy: {metadata.get('max_pleiotropy', 'N/A')}"])
        rows.append(["#"])
        
        # Add header row
        header = [
            "Gene_Symbol", "Gene_Description", "Entrez_ID", "Pleiotropy_Score", 
            "Specificity_Score", "Evidence_Quality", "Literature_Score", 
            "Conservation_Score", "Composite_Score", "BP_Term_Count",
            "Experimental_Evidence_Count", "Computational_Evidence_Count", 
            "IEA_Evidence_Count", "KEGG_Pathway_Count", "PubMed_Count",
            "Guide_Sequence", "Guide_Position", "Doench_Score", 
            "CFD_Off_Targets", "Pathway_Conflict",
            "Phenotype_Severity", "Phenotype_Risk_Level", "Phenotype_Confidence",
            "Phenotype_Lethality_Stage", "Safety_Level", "Safety_Recommendation",
            "Safety_Justification"
        ]
        rows.append(header)
        
        # Add data rows
        for gene in genes:
            gene_symbol = gene.get("symbol", "")
            gene_description = gene.get("description", "")
            entrez_id = gene.get("entrez_id", "")
            pleiotropy_score = gene.get("pleiotropy_score", "")
            specificity_score = gene.get("specificity_score", "")
            evidence_quality = gene.get("evidence_quality", "")
            literature_score = gene.get("literature_score", "")
            conservation_score = gene.get("conservation_score", "")
            composite_score = gene.get("composite_score", "")
            bp_term_count = gene.get("bp_term_count", "")
            exp_count = gene.get("experimental_evidence_count", "")
            comp_count = gene.get("computational_evidence_count", "")
            iea_count = gene.get("iea_evidence_count", "")
            kegg_count = gene.get("kegg_pathway_count", "")
            pubmed_count = gene.get("pubmed_count", "")
            
            # Get phenotype prediction if available
            phenotype_pred = gene.get("phenotype_prediction", {}) or {}
            phenotype_severity = _extract_enum_value(phenotype_pred.get("severity", ""))
            phenotype_risk = _extract_enum_value(phenotype_pred.get("risk_level", ""))
            phenotype_confidence = phenotype_pred.get("confidence_score", "")
            phenotype_stage = phenotype_pred.get("lethality_stage", "")
            
            # Get safety recommendation
            safety_rec = gene.get("safety_recommendation", {}) or {}
            safety_level = safety_rec.get("safety_level", "")
            primary_rec = safety_rec.get("primary_recommendation", "")
            safety_justification = safety_rec.get("justification", "")
            
            # Add rows for each guide
            guides = gene.get("guides", [])
            for guide in guides:
                guide_seq = guide.get("seq", "") if isinstance(guide, dict) else ""
                guide_position = guide.get("position", "") if isinstance(guide, dict) else ""
                doench_score = guide.get("doench_score", "") if isinstance(guide, dict) else ""
                cfd_off_targets = guide.get("cfd_off_targets", "") if isinstance(guide, dict) else ""
                pathway_conflict = guide.get("pathway_conflict", "") if isinstance(guide, dict) else ""
                
                row = [
                    gene_symbol, gene_description, entrez_id, pleiotropy_score,
                    specificity_score, evidence_quality, literature_score,
                    conservation_score, composite_score, bp_term_count,
                    exp_count, comp_count, iea_count, kegg_count, pubmed_count,
                    guide_seq, guide_position, doench_score,
                    cfd_off_targets, pathway_conflict,
                    phenotype_severity, phenotype_risk, phenotype_confidence,
                    phenotype_stage, safety_level, primary_rec, safety_justification
                ]
                rows.append(row)
            
            # If no guides, still add the gene information
            if not guides:
                row = [
                    gene_symbol, gene_description, entrez_id, pleiotropy_score,
                    specificity_score, evidence_quality, literature_score,
                    conservation_score, composite_score, bp_term_count,
                    exp_count, comp_count, iea_count, kegg_count, pubmed_count,
                    "", "", "", "", "",
                    phenotype_severity, phenotype_risk, phenotype_confidence,
                    phenotype_stage, safety_level, primary_rec, safety_justification
                ]
                rows.append(row)
        
        # Write CSV file
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(rows)
        
        logger.info(f"Generated CSV report at {csv_path}")
        
    except Exception as e:
        logger.error(f"Failed to generate CSV report: {str(e)}")
        raise


def _generate_gene_summary_csv(pipeline_output: dict, csv_path: str) -> None:
    """
    Generate a gene-focused summary CSV report.
    
    Args:
        pipeline_output: Output dictionary from run_k_sites_pipeline
        csv_path: Path to save the CSV report
    """
    try:
        genes = pipeline_output.get("genes", [])
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'gene_symbol', 'gene_description', 'entrez_id', 
                'pleiotropy_score', 'specificity_score', 'composite_score',
                'evidence_quality', 'literature_score', 'conservation_score',
                'bp_term_count', 'experimental_evidence_count', 
                'computational_evidence_count', 'iea_evidence_count',
                'kegg_pathway_count', 'pubmed_count', 'guide_count',
                'phenotype_severity', 'phenotype_risk_level', 
                'safety_level', 'safety_recommendation', 'safety_justification'
            ]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for gene in genes:
                phenotype_pred = gene.get("phenotype_prediction", {}) or {}
                safety_rec = gene.get("safety_recommendation", {}) or {}
                
                writer.writerow({
                    'gene_symbol': gene.get('symbol', ''),
                    'gene_description': gene.get('description', ''),
                    'entrez_id': gene.get('entrez_id', ''),
                    'pleiotropy_score': gene.get('pleiotropy_score', ''),
                    'specificity_score': gene.get('specificity_score', ''),
                    'composite_score': gene.get('composite_score', ''),
                    'evidence_quality': gene.get('evidence_quality', ''),
                    'literature_score': gene.get('literature_score', ''),
                    'conservation_score': gene.get('conservation_score', ''),
                    'bp_term_count': gene.get('bp_term_count', ''),
                    'experimental_evidence_count': gene.get('experimental_evidence_count', ''),
                    'computational_evidence_count': gene.get('computational_evidence_count', ''),
                    'iea_evidence_count': gene.get('iea_evidence_count', ''),
                    'kegg_pathway_count': gene.get('kegg_pathway_count', ''),
                    'pubmed_count': gene.get('pubmed_count', ''),
                    'guide_count': len(gene.get('guides', [])),
                    'phenotype_severity': _extract_enum_value(phenotype_pred.get('severity', '')),
                    'phenotype_risk_level': _extract_enum_value(phenotype_pred.get('risk_level', '')),
                    'safety_level': safety_rec.get('safety_level', ''),
                    'safety_recommendation': safety_rec.get('primary_recommendation', ''),
                    'safety_justification': safety_rec.get('justification', '')
                })
        
        logger.info(f"Generated gene summary CSV at {csv_path}")
        
    except Exception as e:
        logger.error(f"Failed to generate gene summary CSV: {str(e)}")


def _generate_excel_report(pipeline_output: dict, excel_path: str) -> None:
    """
    Generate a comprehensive Excel report with multiple sheets.
    
    Args:
        pipeline_output: Output dictionary from run_k_sites_pipeline
        excel_path: Path to save the Excel report
    """
    try:
        # Try to import openpyxl for Excel generation
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.warning("openpyxl not available, generating CSV fallback instead of Excel")
            # Generate CSV as fallback
            csv_fallback = excel_path.replace('.xlsx', '_excel_fallback.csv')
            _generate_csv_report(pipeline_output, csv_fallback)
            return
        
        genes = pipeline_output.get("genes", [])
        metadata = pipeline_output.get("metadata", {})
        statistics = pipeline_output.get("statistics", {})
        
        wb = Workbook()
        
        # ===== Sheet 1: Summary =====
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # Add metadata
        ws_summary['A1'] = "K-Sites CRISPR Design Report"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A3'] = "GO Term:"
        ws_summary['B3'] = metadata.get('go_term', 'N/A')
        ws_summary['A4'] = "Organism:"
        ws_summary['B4'] = metadata.get('organism', 'N/A')
        ws_summary['A5'] = "Timestamp:"
        ws_summary['B5'] = metadata.get('timestamp', 'N/A')
        ws_summary['A6'] = "Evidence Filter:"
        ws_summary['B6'] = metadata.get('evidence_filter', 'N/A')
        ws_summary['A7'] = "Max Pleiotropy:"
        ws_summary['B7'] = metadata.get('max_pleiotropy', 'N/A')
        
        # Add statistics
        ws_summary['A9'] = "Statistics"
        ws_summary['A9'].font = Font(bold=True, size=12)
        ws_summary['A10'] = "Total Genes Screened:"
        ws_summary['B10'] = statistics.get('total_genes_screened', 0)
        ws_summary['A11'] = "Genes Passed Filter:"
        ws_summary['B11'] = statistics.get('genes_passed_filter', 0)
        ws_summary['A12'] = "Average Pleiotropy:"
        ws_summary['B12'] = f"{statistics.get('avg_pleiotropy', 0):.2f}"
        
        # ===== Sheet 2: Gene Rankings =====
        ws_genes = wb.create_sheet("Gene Rankings")
        
        # Headers
        gene_headers = [
            "Rank", "Gene Symbol", "Pleiotropy Score", "Specificity Score",
            "Composite Score", "Evidence Quality", "Literature Score",
            "Conservation Score", "BP Term Count", "Safety Level"
        ]
        
        for col, header in enumerate(gene_headers, 1):
            cell = ws_genes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data
        for row, gene in enumerate(genes, 2):
            ws_genes.cell(row=row, column=1, value=row-1)  # Rank
            ws_genes.cell(row=row, column=2, value=gene.get('symbol', ''))
            ws_genes.cell(row=row, column=3, value=gene.get('pleiotropy_score', ''))
            ws_genes.cell(row=row, column=4, value=gene.get('specificity_score', ''))
            ws_genes.cell(row=row, column=5, value=gene.get('composite_score', ''))
            ws_genes.cell(row=row, column=6, value=gene.get('evidence_quality', ''))
            ws_genes.cell(row=row, column=7, value=gene.get('literature_score', ''))
            ws_genes.cell(row=row, column=8, value=gene.get('conservation_score', ''))
            ws_genes.cell(row=row, column=9, value=gene.get('bp_term_count', ''))
            
            safety_rec = gene.get('safety_recommendation', {}) or {}
            ws_genes.cell(row=row, column=10, value=safety_rec.get('safety_level', ''))
        
        # ===== Sheet 3: gRNA Designs =====
        ws_guides = wb.create_sheet("gRNA Designs")
        
        guide_headers = [
            "Gene Symbol", "gRNA Sequence", "Position", "Doench Score",
            "CFD Off-targets", "Pathway Conflict"
        ]
        
        for col, header in enumerate(guide_headers, 1):
            cell = ws_guides.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        guide_row = 2
        for gene in genes:
            gene_symbol = gene.get('symbol', '')
            for guide in gene.get('guides', []):
                if isinstance(guide, dict):
                    ws_guides.cell(row=guide_row, column=1, value=gene_symbol)
                    ws_guides.cell(row=guide_row, column=2, value=guide.get('seq', ''))
                    ws_guides.cell(row=guide_row, column=3, value=guide.get('position', ''))
                    ws_guides.cell(row=guide_row, column=4, value=guide.get('doench_score', ''))
                    ws_guides.cell(row=guide_row, column=5, value=guide.get('cfd_off_targets', ''))
                    ws_guides.cell(row=guide_row, column=6, value=guide.get('pathway_conflict', False))
                    guide_row += 1
        
        # ===== Sheet 4: Safety Recommendations =====
        ws_safety = wb.create_sheet("Safety Recommendations")
        
        safety_headers = [
            "Gene Symbol", "Safety Level", "Primary Recommendation",
            "Justification", "Concerns", "Mitigation Strategies"
        ]
        
        for col, header in enumerate(safety_headers, 1):
            cell = ws_safety.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, gene in enumerate(genes, 2):
            safety_rec = gene.get('safety_recommendation', {}) or {}
            ws_safety.cell(row=row, column=1, value=gene.get('symbol', ''))
            ws_safety.cell(row=row, column=2, value=safety_rec.get('safety_level', ''))
            ws_safety.cell(row=row, column=3, value=safety_rec.get('primary_recommendation', ''))
            ws_safety.cell(row=row, column=4, value=safety_rec.get('justification', ''))
            ws_safety.cell(row=row, column=5, value='; '.join(safety_rec.get('concerns', [])))
            ws_safety.cell(row=row, column=6, value='; '.join(safety_rec.get('mitigation_strategies', [])))
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_genes, ws_guides, ws_safety]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(excel_path)
        logger.info(f"Generated Excel report at {excel_path}")
        
    except Exception as e:
        logger.error(f"Failed to generate Excel report: {str(e)}")


def _generate_fasta_report(pipeline_output: dict, fasta_path: str) -> None:
    """
    Generate a FASTA file with all gRNA sequences.
    
    Args:
        pipeline_output: Output dictionary from run_k_sites_pipeline
        fasta_path: Path to save the FASTA report
    """
    try:
        genes = pipeline_output.get("genes", [])
        metadata = pipeline_output.get("metadata", {})
        
        with open(fasta_path, 'w', encoding='utf-8') as fasta_file:
            # Write header comment
            fasta_file.write(f"; K-Sites gRNA sequences\n")
            fasta_file.write(f"; GO Term: {metadata.get('go_term', 'N/A')}\n")
            fasta_file.write(f"; Organism: {metadata.get('organism', 'N/A')}\n")
            fasta_file.write(f"; Generated from K-Sites pipeline\n\n")
            
            # Write sequences
            for gene in genes:
                gene_symbol = gene.get("symbol", "Unknown")
                guides = gene.get("guides", [])
                
                for i, guide in enumerate(guides):
                    if isinstance(guide, dict):
                        guide_seq = guide.get("seq", "")
                        if guide_seq:
                            # Create header with gene info and guide properties
                            header_parts = [
                                f"gene={gene_symbol}",
                                f"pos={guide.get('position', 'N/A')}",
                                f"doench={guide.get('doench_score', 'N/A'):.2f}" if isinstance(guide.get('doench_score'), (int, float)) else f"doench={guide.get('doench_score', 'N/A')}"
                            ]
                            if guide.get("pathway_conflict", False):
                                header_parts.append("pathway_conflict=YES")
                            
                            header_info = " ".join(header_parts)
                            fasta_file.write(f">gRNA_{gene_symbol}_{i+1} {header_info}\n")
                            fasta_file.write(f"{guide_seq}\n")
        
        logger.info(f"Generated FASTA report at {fasta_path}")
        
    except Exception as e:
        logger.error(f"Failed to generate FASTA report: {str(e)}")


def _extract_enum_value(value) -> str:
    """Extract string value from enum or return as-is if string."""
    if hasattr(value, 'value'):
        return str(value.value)
    return str(value) if value else ""


def _generate_report_html(
    metadata: Dict[str, Any], 
    genes: list, 
    total_genes_screened: int, 
    total_guides: int, 
    pathway_conflict_guides: int,
    output_path: str
) -> str:
    """
    Generate the complete HTML report.
    """
    # Escape metadata values for safety
    escaped_go_term = html.escape(metadata.get("go_term", "Unknown"))
    escaped_organism = html.escape(metadata.get("organism", "Unknown"))
    resolved_organism = metadata.get("resolved_organism", {})
    escaped_scientific_name = html.escape(resolved_organism.get("scientific_name", "Unknown") if resolved_organism else "Unknown")
    timestamp = metadata.get("timestamp", "")
    execution_duration = metadata.get("execution_duration", 0)
    evidence_filter = metadata.get("evidence_filter", "experimental")
    max_pleiotropy = metadata.get("max_pleiotropy", 10)
    
    # Build the HTML content
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>K-Sites CRISPR Design Report - {escaped_go_term}</title>
    <style>
        :root {{
            --primary-color: #2c3e50;
            --secondary-color: #3498db;
            --success-color: #27ae60;
            --warning-color: #f39c12;
            --danger-color: #e74c3c;
            --light-bg: #f8f9fa;
            --border-color: #dee2e6;
            --text-color: #212529;
            --text-light: #6c757d;
        }}
        
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            line-height: 1.6;
            color: var(--text-color);
            background-color: white;
            padding: 20px;
        }}
        
        .container {{ max-width: 1200px; margin: 0 auto; }}
        
        header {{
            text-align: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 2px solid var(--border-color);
        }}
        
        h1 {{ color: var(--primary-color); margin-bottom: 10px; font-size: 2.2em; }}
        
        .subtitle {{ color: var(--text-light); font-size: 1.1em; margin-bottom: 15px; }}
        
        .summary-card {{
            background: linear-gradient(135deg, #f5f7fa 0%, #e4edf9 100%);
            border-radius: 10px;
            padding: 20px;
            margin: 20px 0;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            border-left: 4px solid var(--secondary-color);
        }}
        
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-top: 15px;
        }}
        
        .summary-item {{
            text-align: center;
            padding: 15px;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }}
        
        .summary-value {{ font-size: 1.8em; font-weight: bold; color: var(--primary-color); }}
        .summary-label {{ font-size: 0.9em; color: var(--text-light); }}
        
        .section {{ margin: 30px 0; }}
        
        .section-title {{
            color: var(--primary-color);
            border-bottom: 2px solid var(--secondary-color);
            padding-bottom: 10px;
            margin-bottom: 20px;
            font-size: 1.5em;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
            background: white;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            border-radius: 8px;
            overflow: hidden;
        }}
        
        th, td {{ padding: 12px 15px; text-align: left; border-bottom: 1px solid var(--border-color); }}
        th {{ background-color: var(--primary-color); color: white; font-weight: 600; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        tr:hover {{ background-color: #e9f7fe; }}
        
        .sequence {{
            font-family: 'SF Mono', Consolas, Monaco, monospace;
            font-size: 0.9em;
            letter-spacing: 0.5px;
            background-color: #f8f9fa;
            padding: 2px 6px;
            border-radius: 4px;
            border: 1px solid #e9ecef;
        }}
        
        .score-high {{ color: var(--success-color); font-weight: bold; }}
        .score-medium {{ color: var(--warning-color); font-weight: bold; }}
        .score-low {{ color: var(--danger-color); font-weight: bold; }}
        
        .pathway-conflict {{
            background-color: #fff3cd;
            color: #856404;
            padding: 2px 6px;
            border-radius: 4px;
            font-weight: bold;
        }}
        
        .phenotype-severity, .risk-level, .safety-level {{
            display: inline-block;
            padding: 4px 8px;
            border-radius: 12px;
            font-size: 0.8em;
            font-weight: bold;
            text-transform: uppercase;
        }}
        
        .safety-safe {{ background-color: #d4edda; color: #155724; }}
        .safety-caution {{ background-color: #fff3cd; color: #856404; }}
        .safety-warning {{ background-color: #f8d7da; color: #721c24; }}
        .safety-critical {{ background-color: #dc3545; color: white; }}
        
        .risk-critical {{ background-color: #dc3545; color: white; }}
        .risk-high {{ background-color: #fd7e14; color: white; }}
        .risk-medium {{ background-color: #ffc107; color: #212529; }}
        .risk-low {{ background-color: #28a745; color: white; }}
        .risk-unknown {{ background-color: #6c757d; color: white; }}
        
        .visual-indicator {{
            display: inline-block;
            width: 12px;
            height: 12px;
            border-radius: 50%;
            margin-right: 5px;
        }}
        
        .high-specificity {{ background-color: #28a745; }}
        .medium-specificity {{ background-color: #ffc107; }}
        .low-specificity {{ background-color: #dc3545; }}
        
        .specificity-bar {{
            height: 10px;
            background-color: #e9ecef;
            border-radius: 5px;
            overflow: hidden;
            margin: 5px 0;
            width: 100px;
        }}
        
        .specificity-fill {{ height: 100%; background-color: #28a745; }}
        
        .btn-copy {{
            background-color: #e9ecef;
            border: none;
            padding: 4px 8px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 0.8em;
        }}
        .btn-copy:hover {{ background-color: #dee2e6; }}
        
        .download-section {{
            text-align: center;
            margin: 40px 0;
            padding: 30px;
            background: linear-gradient(135deg, #e8f4f8 0%, #d4edf9 100%);
            border-radius: 10px;
        }}
        
        .download-btn {{
            display: inline-block;
            background-color: var(--secondary-color);
            color: white;
            padding: 12px 24px;
            text-decoration: none;
            border-radius: 8px;
            font-weight: 600;
            margin: 10px;
            transition: all 0.3s;
        }}
        .download-btn:hover {{
            background-color: #2980b9;
            transform: translateY(-2px);
        }}
        
        .accordion {{
            background-color: white;
            border: 1px solid var(--border-color);
            border-radius: 8px;
            margin-bottom: 10px;
            overflow: hidden;
        }}
        
        .accordion-header {{
            background-color: #f8f9fa;
            padding: 15px;
            cursor: pointer;
            display: flex;
            justify-content: space-between;
            align-items: center;
            font-weight: 600;
        }}
        
        .accordion-content {{ padding: 0; max-height: 0; overflow: hidden; transition: max-height 0.2s ease-out; }}
        .accordion.active .accordion-content {{ max-height: 2000px; }}
        .accordion-body {{ padding: 20px; }}
        
        .justification-box {{
            background-color: #f8f9fa;
            padding: 10px;
            border-radius: 4px;
            font-size: 0.9em;
            margin-top: 5px;
            border-left: 3px solid var(--secondary-color);
        }}
        
        @media (max-width: 768px) {{
            body {{ padding: 10px; }}
            .summary-grid {{ grid-template-columns: 1fr; }}
            th, td {{ padding: 8px 10px; font-size: 0.9em; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>🧬 K-Sites CRISPR Design Report</h1>
            <div class="subtitle">
                GO Term: <strong>{escaped_go_term}</strong> | 
                Organism: <strong>{escaped_scientific_name}</strong> | 
                Generated: {timestamp}
            </div>
            <div class="subtitle">
                Evidence Filter: <strong>{evidence_filter}</strong> | 
                Max Pleiotropy: <strong>{max_pleiotropy}</strong>
            </div>
        </header>
        
        <div class="summary-card">
            <h2>📊 Executive Summary</h2>
            <div class="summary-grid">
                <div class="summary-item">
                    <div class="summary-value">{total_genes_screened}</div>
                    <div class="summary-label">Genes Screened</div>
                </div>
                <div class="summary-item">
                    <div class="summary-value">{total_guides}</div>
                    <div class="summary-label">gRNAs Designed</div>
                </div>
                <div class="summary-item">
                    <div class="summary-value">{len([g for gene in genes for g in gene.get('guides', []) if isinstance(g, dict) and g.get('doench_score', 0) >= 0.7])}</div>
                    <div class="summary-label">High-Efficacy gRNAs</div>
                </div>
                <div class="summary-item">
                    <div class="summary-value">{pathway_conflict_guides}</div>
                    <div class="summary-label">Pathway Conflicts</div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2 class="section-title">🧬 Gene Rankings with Specificity Indicators</h2>
            <p style="margin-bottom: 15px; color: var(--text-light);">
                Genes ranked by composite score combining specificity (40%), evidence quality (25%), 
                literature support (20%), and conservation (15%). Specificity score is on 0-1 scale 
                (higher = more specific, less pleiotropic).
            </p>
            <table>
                <thead>
                    <tr>
                        <th>Rank</th>
                        <th>Gene Symbol</th>
                        <th>Pleiotropy (0-10)</th>
                        <th>Specificity (0-1)</th>
                        <th>Visual</th>
                        <th>Composite Score</th>
                        <th>Safety Level</th>
                        <th>Recommendation</th>
                        <th>gRNAs</th>
                    </tr>
                </thead>
                <tbody>
"""
    
    # Add gene rows to the table
    for rank, gene in enumerate(genes, 1):
        symbol = html.escape(gene.get("symbol", "Unknown"))
        pleiotropy_score = gene.get("pleiotropy_score", 0)
        specificity_score = gene.get("specificity_score", 0)
        composite_score = gene.get("composite_score", 0)
        
        # Ensure specificity is 0-1 scale
        if specificity_score > 1.0:
            specificity_score = specificity_score / 10.0
        
        specificity_percentage = specificity_score * 100
        
        # Get safety recommendation
        safety_rec = gene.get("safety_recommendation", {}) or {}
        safety_level = safety_rec.get("safety_level", "UNKNOWN")
        primary_rec = safety_rec.get("primary_recommendation", "")
        safety_justification = safety_rec.get("justification", "")
        
        # Determine visual indicator class
        if specificity_score >= 0.7:
            indicator_class = "high-specificity"
        elif specificity_score >= 0.4:
            indicator_class = "medium-specificity"
        else:
            indicator_class = "low-specificity"
        
        # Determine safety class
        safety_class = "safety-safe"
        if safety_level in ["CRITICAL", "HIGH_RISK"]:
            safety_class = "safety-critical"
        elif safety_level in ["WARNING", "CAUTION"]:
            safety_class = "safety-warning"
        elif safety_level == "MODERATE":
            safety_class = "safety-caution"
        
        guide_count = len(gene.get("guides", []))
        
        html_content += f"""
                    <tr>
                        <td>{rank}</td>
                        <td><strong>{symbol}</strong></td>
                        <td>{pleiotropy_score:.2f}</td>
                        <td>{specificity_score:.3f}</td>
                        <td>
                            <div class="visual-indicator {indicator_class}"></div>
                            <div class="specificity-bar">
                                <div class="specificity-fill" style="width: {specificity_percentage}%"></div>
                            </div>
                        </td>
                        <td>{composite_score:.3f}</td>
                        <td><span class="safety-level {safety_class}">{safety_level}</span></td>
                        <td>
                            {primary_rec}
                            {f'<div class="justification-box">{html.escape(safety_justification)}</div>' if safety_justification else ''}
                        </td>
                        <td>{guide_count}</td>
                    </tr>
"""
    
    html_content += """
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2 class="section-title">🔬 Detailed gRNA Designs</h2>
"""
    
    # Add detailed gRNA information for each gene
    for gene in genes:
        symbol = html.escape(gene.get("symbol", "Unknown"))
        guides = gene.get("guides", [])
        
        if not guides:
            continue
            
        html_content += f"""
            <div class="accordion">
                <div class="accordion-header" onclick="toggleAccordion(this)">
                    <span>{symbol} - {len(guides)} gRNA{'s' if len(guides) != 1 else ''}</span>
                    <span>▼</span>
                </div>
                <div class="accordion-content">
                    <div class="accordion-body">
                        <table>
                            <thead>
                                <tr>
                                    <th>gRNA Sequence (5'→3')</th>
                                    <th>Position</th>
                                    <th>Doench Score</th>
                                    <th>CFD Off-targets</th>
                                    <th>Pathway Conflict</th>
                                    <th>Action</th>
                                </tr>
                            </thead>
                            <tbody>
"""
        
        for guide in guides:
            if isinstance(guide, dict):
                seq = html.escape(guide.get("seq", ""))
                position = guide.get("position", "N/A")
                doench_score = guide.get("doench_score", 0)
                cfd_off_targets = guide.get("cfd_off_targets", "N/A")
                pathway_conflict = guide.get("pathway_conflict", False)
                
                # Determine score class
                if isinstance(doench_score, (int, float)):
                    if doench_score >= 0.7:
                        score_class = "score-high"
                    elif doench_score >= 0.5:
                        score_class = "score-medium"
                    else:
                        score_class = "score-low"
                    doench_display = f"{doench_score:.2f}"
                else:
                    score_class = ""
                    doench_display = str(doench_score)
                
                conflict_indicator = '<span class="pathway-conflict">⚠️ Yes</span>' if pathway_conflict else "No"
                
                html_content += f"""
                                <tr>
                                    <td><span class="sequence">{seq}</span></td>
                                    <td>{position}</td>
                                    <td class="{score_class}">{doench_display}</td>
                                    <td>{cfd_off_targets}</td>
                                    <td>{conflict_indicator}</td>
                                    <td><button class="btn-copy" onclick="copyToClipboard('{seq}')">📋 Copy</button></td>
                                </tr>
"""
        
        html_content += """
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
"""
    
    # Get base filename for download links
    base_name = os.path.basename(output_path).replace('.html', '')
    
    html_content += f"""
        </div>
        
        <div class="download-section">
            <h2>📥 Download Results</h2>
            <p style="margin-bottom: 20px; color: var(--text-light);">
                Download results in various formats for further analysis.
            </p>
            <a href="./{base_name}_comprehensive.csv" class="download-btn">📄 Comprehensive CSV</a>
            <a href="./{base_name}_gene_summary.csv" class="download-btn">📊 Gene Summary CSV</a>
            <a href="./{base_name}_comprehensive.xlsx" class="download-btn">📗 Excel Report</a>
            <a href="./{base_name}_grna_sequences.fasta" class="download-btn">🧬 FASTA Sequences</a>
        </div>
        
        <footer style="text-align: center; margin-top: 40px; padding-top: 20px; border-top: 1px solid var(--border-color); color: var(--text-light); font-size: 0.9em;">
            <p>K-Sites CRISPR Design Platform | Non-Pleiotropic Gene Identification System</p>
            <p>Execution time: {execution_duration:.2f}s | Evidence: {evidence_filter} | Max Pleiotropy: {max_pleiotropy}</p>
        </footer>
    </div>
    
    <script>
        function toggleAccordion(header) {{
            const accordion = header.parentElement;
            accordion.classList.toggle('active');
            const arrow = header.querySelector('span:last-child');
            arrow.textContent = accordion.classList.contains('active') ? '▲' : '▼';
        }}
        
        function copyToClipboard(text) {{
            navigator.clipboard.writeText(text).then(() => {{
                alert('Copied to clipboard: ' + text);
            }}).catch(err => {{
                const textArea = document.createElement('textarea');
                textArea.value = text;
                document.body.appendChild(textArea);
                textArea.select();
                document.execCommand('copy');
                document.body.removeChild(textArea);
                alert('Copied to clipboard: ' + text);
            }});
        }}
    </script>
</body>
</html>"""
    
    return html_content
